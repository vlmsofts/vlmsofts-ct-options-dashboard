"""
Read live CT market data from the RTD Site Excel workbook — read-only, never modifies anything.

Priority:
  1. win32com  — reads from the running Excel/Bloomberg session (live, intraday)
  2. openpyxl  — reads from the saved xlsx on OneDrive (last AutoSave snapshot)
  3. Returns None — caller falls back to GitHub CSV data

Sheets read:
  BB Futures Data  — outright futures + calendar spreads + currencies
  BB Options Data  — ATM straddle / call / put per expiry
  RTD              — Cotton Straddles summary (ATM vol, strad value, % chg)

Ticker format: Bloomberg always uses 'CTN6 Comdty', 'CTN6CTZ6 Comdty' etc.
The reader strips ' Comdty' to produce dashboard tickers: 'CTN6', 'CTN6CTZ6'.
Rows are read dynamically — any ticker present is captured; none are hardcoded.
"""

import os
import logging

log = logging.getLogger(__name__)

RTD_XLSX_PATH  = r"C:\Users\Louis\OneDrive - VLM Commodities LTD\Site Sync\RTD SIte.xlsx"
WORKBOOK_FRAGMENT = "RTD SIte"   # matches 'RTD SIte.xlsm' or 'RTD SIte.xlsx'

# Bloomberg error strings that mean "no data" — treated as None throughout
_BBG_ERRORS = {
    '#n/a review', '#value!', '#ref!', '#n/a', '#div/0!',
    '#name?', '#null!', '#num!', 'buh bye',
}


# ── Public entry point ────────────────────────────────────────────────────────

def read_sheet_sample(sheet_fragment, max_rows=5):
    """Return first max_rows rows from the sheet matching sheet_fragment, for diagnostics."""
    try:
        import win32com.client
        xl = win32com.client.GetActiveObject("Excel.Application")
        for i in range(1, xl.Workbooks.Count + 1):
            if WORKBOOK_FRAGMENT.lower() in xl.Workbooks(i).Name.lower():
                wb = xl.Workbooks(i)
                for j in range(1, wb.Sheets.Count + 1):
                    name = wb.Sheets(j).Name
                    if sheet_fragment.lower() in name.lower():
                        ws   = wb.Sheets(j)
                        used = ws.UsedRange.Value
                        if used is None:
                            return {'sheet': name, 'rows': [], 'note': 'UsedRange.Value is None'}
                        if not isinstance(used, tuple):
                            return {'sheet': name, 'rows': [[str(used)]], 'note': 'single cell'}
                        if not isinstance(used[0], tuple):
                            rows = [list(used)]
                        else:
                            rows = [list(r) for r in used]
                        sample = [[str(v) for v in row] for row in rows[:max_rows]]
                        return {'sheet': name, 'total_rows': len(rows),
                                'total_cols': len(rows[0]) if rows else 0,
                                'sample': sample}
        return {'error': 'sheet not found'}
    except Exception as e:
        return {'error': str(e)}


def workbook_sheet_names():
    """Return list of sheet names from the open RTD workbook, or [] if unavailable."""
    try:
        import win32com.client
        xl = win32com.client.GetActiveObject("Excel.Application")
        for i in range(1, xl.Workbooks.Count + 1):
            if WORKBOOK_FRAGMENT.lower() in xl.Workbooks(i).Name.lower():
                wb = xl.Workbooks(i)
                return [wb.Sheets(j).Name for j in range(1, wb.Sheets.Count + 1)]
    except Exception:
        pass
    try:
        import openpyxl
        if os.path.exists(RTD_XLSX_PATH):
            wb = openpyxl.load_workbook(RTD_XLSX_PATH, read_only=True, data_only=True)
            names = wb.sheetnames
            wb.close()
            return names
    except Exception:
        pass
    return []


def read_live():
    """
    Return dict of live market data, or None if unavailable.
    Keys: 'outrights', 'spreads', 'options', 'straddles', 'source'
    """
    for reader, label in [(_read_via_com, 'live_rtd'), (_read_via_openpyxl, 'saved_xlsx')]:
        try:
            result = reader()
            if result:
                result['source'] = label
                log.info(f"RTD data loaded via {label}")
                return result
        except Exception as e:
            log.debug(f"RTD {label} read failed: {e}")

    log.debug("RTD unavailable — using CSV fallback")
    return None


# ── win32com reader ────────────────────────────────────────────────────────────

def _read_via_com():
    import win32com.client
    import pythoncom

    # Flask serves requests on worker threads; COM requires each thread to initialise
    # its own apartment before calling GetActiveObject, otherwise the call raises
    # CO_E_NOTINITIALIZED and we silently fall through to the openpyxl (XLSX) path.
    pythoncom.CoInitialize()
    try:
        return _read_via_com_inner()
    finally:
        pythoncom.CoUninitialize()


def _read_via_com_inner():
    import win32com.client

    xl = win32com.client.GetActiveObject("Excel.Application")

    wb = None
    for i in range(1, xl.Workbooks.Count + 1):
        if WORKBOOK_FRAGMENT.lower() in xl.Workbooks(i).Name.lower():
            wb = xl.Workbooks(i)
            break
    if wb is None:
        raise RuntimeError("RTD workbook not open in Excel")

    def sheet_rows(name):
        ws   = wb.Sheets(name)
        used = ws.UsedRange.Value      # bulk read — fast
        if used is None:
            return []
        if not isinstance(used, tuple):
            return [[used]]
        if not isinstance(used[0], tuple):
            return [list(used)]
        return [list(row) for row in used]

    def find_all_options_sheet():
        """Find the 'all options' BQL sheet by name or by content signature."""
        for i in range(1, wb.Sheets.Count + 1):
            name = wb.Sheets(i).Name
            if 'all options' in name.lower():
                try:
                    return sheet_rows(name)
                except Exception:
                    pass
        # Name not found — scan each sheet for the security_des content signature
        for i in range(1, wb.Sheets.Count + 1):
            try:
                rows = sheet_rows(wb.Sheets(i).Name)
                if rows and rows[0] and any(
                    v is not None and 'security_des' in str(v).lower()
                    for v in rows[0]
                ):
                    return rows
            except Exception:
                continue
        return []

    return _parse_all(
        sheet_rows("BB Futures Data"),
        sheet_rows("BB Options Data"),
        sheet_rows("RTD"),
        find_all_options_sheet(),
    )


# ── openpyxl reader ────────────────────────────────────────────────────────────

def _read_via_openpyxl():
    import openpyxl

    if not os.path.exists(RTD_XLSX_PATH):
        raise FileNotFoundError(RTD_XLSX_PATH)

    wb = openpyxl.load_workbook(RTD_XLSX_PATH, read_only=True, data_only=True)
    try:
        def sheet_rows(name):
            return [list(row) for row in wb[name].iter_rows(values_only=True)]

        def find_all_options_sheet():
            for name in wb.sheetnames:
                if 'all options' in name.lower():
                    try:
                        return sheet_rows(name)
                    except Exception:
                        pass
            # Scan by content signature
            for name in wb.sheetnames:
                try:
                    rows = sheet_rows(name)
                    if rows and rows[0] and any(
                        v is not None and 'security_des' in str(v).lower()
                        for v in rows[0]
                    ):
                        return rows
                except Exception:
                    continue
            return []

        return _parse_all(
            sheet_rows("BB Futures Data"),
            sheet_rows("BB Options Data"),
            sheet_rows("RTD"),
            find_all_options_sheet(),
        )
    finally:
        wb.close()


# ── Master parser ─────────────────────────────────────────────────────────────

def _parse_all(fut_rows, opt_rows, rtd_rows, all_opt_rows=None):
    outrights, spreads = _parse_futures(fut_rows)
    options            = _parse_options(opt_rows)
    straddles          = _parse_rtd(rtd_rows)
    all_parsed         = _parse_all_options(all_opt_rows or [])
    live_options       = all_parsed['options']
    live_futures_sheet = all_parsed['futures']

    # Merge live futures from 'all options' sheet into outrights
    for tkr, lf in live_futures_sheet.items():
        if tkr in outrights:
            if lf.get('last') and lf['last'] > 0:
                outrights[tkr]['last'] = lf['last']
            if lf.get('settle') and lf['settle'] > 0:
                outrights[tkr]['settle'] = lf['settle']
        elif lf.get('last') or lf.get('settle'):
            outrights[tkr] = {
                'ticker': tkr, 'last': lf.get('last'), 'settle': lf.get('settle'),
                'change': None, 'oi': None, 'volume': None,
            }

    # Merge RTD straddle data into options dict (keyed by dashboard ticker e.g. 'CTN6')
    for contract_code, sd in straddles.items():
        ticker = 'CT' + contract_code          # 'N6' -> 'CTN6'
        if ticker in options:
            if sd.get('atm_vol') is not None:
                options[ticker]['atm_vol_rtd']      = sd['atm_vol']
            if sd.get('value') is not None:
                options[ticker]['strad_val_rtd']    = sd['value']
            if sd.get('pct_chg') is not None:
                options[ticker]['pct_chg_rtd']      = sd['pct_chg']
            if sd.get('dte') is not None:
                options[ticker]['dte_rtd']          = sd['dte']
            if sd.get('settlement') is not None:
                options[ticker]['strad_settle_rtd'] = sd['settlement']
            if sd.get('strike') is not None:
                options[ticker]['atm_strike_rtd']   = sd['strike']

    if not outrights and not options:
        return None

    return {
        'outrights':     outrights,     # {ticker: {...}}
        'spreads':       spreads,       # {ticker: {...}}
        'options':       options,       # {ticker: {...}}
        'straddles':     straddles,     # {contract_code: {...}}
        'live_options':  live_options,  # {ticker: {'expiry': 'YYYY-MM-DD', 'strikes': [...]}}
    }


# ── BB Futures Data parser ────────────────────────────────────────────────────

def _parse_futures(rows):
    """
    Reads every CT row regardless of what tickers are present.
    Classifies by Contract column: 'Jul26' = outright, 'Jul26/Dec26' = spread.
    Skips: blank rows, non-CT tickers, index rows (start with '.'), formula fragments.
    """
    if not rows or not rows[0]:
        return {}, {}

    headers = [_hdr(h) for h in rows[0]]
    col     = _col_map(headers)   # keeps first occurrence of each name

    outrights, spreads = {}, {}

    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue

        raw_ticker = _str(row, col.get('Ticker', 0))
        if not raw_ticker:
            continue
        # Skip non-CT rows, index symbols, and stray formula fragments
        if 'CT' not in raw_ticker or raw_ticker.startswith('.') or raw_ticker.startswith('='):
            continue
        if 'Curncy' in raw_ticker or 'Comdty' not in raw_ticker:
            continue

        ticker   = raw_ticker.replace(' Comdty', '').strip()
        contract = _str(row, col.get('Contract', 1)) or ''

        entry = {
            'ticker':   ticker,
            'contract': contract,
            'last':     _num(row, col.get('Last',        2)),
            'settle':   _num(row, col.get('SettlePrice', 3)),
            'yest':     _num(row, col.get('Yest Settle', 4)),
            'change':   _num(row, col.get('Change',      5)),
            'pct_chg':  _num(row, col.get('% Change',    6)),
            'oi':       _num(row, col.get('Open Int',     7)),
            'oi_chg':   _num(row, col.get('O.I. Chg',    8)),
            'volume':   _num(row, col.get('Volume',       9)),
            'high':     _num(row, col.get('High',        13)),
            'low':      _num(row, col.get('Low',         14)),
            'hv10':     _hv(row,  col.get('HV 10day',   15)),
            'hv30':     _hv(row,  col.get('HV 30day',   16)),
            'hv60':     _hv(row,  col.get('HV 60day',   17)),
            'hv90':     _hv(row,  col.get('HV 90day',   18)),
        }

        # Spread if contract contains '/', otherwise outright future
        if '/' in contract:
            legs = [l.strip() for l in contract.split('/')]
            entry['legs'] = legs
            spreads[ticker] = entry
        else:
            outrights[ticker] = entry

    return outrights, spreads


# ── BB Options Data parser ────────────────────────────────────────────────────

def _parse_options(rows):
    """
    Reads all CT option rows dynamically.
    Stops when it hits the 'Holidays' section.
    The sheet has two columns both named 'Ticker':
      col B (index 1) = contract month label  e.g. 'Jul26'
      col D (index 3) = Bloomberg ticker      e.g. 'CTN6 Comdty'
    Uses Bloomberg ticker (col D) as the key, stripped to 'CTN6'.
    """
    if not rows or not rows[0]:
        return {}

    headers   = [_hdr(h) for h in rows[0]]
    col_all   = {}                          # name -> [list of indices]
    for i, h in enumerate(headers):
        if h:
            col_all.setdefault(h, []).append(i)

    def col(name, occurrence=0):
        idxs = col_all.get(name, [])
        return idxs[occurrence] if occurrence < len(idxs) else None

    # 'Ticker' column appears twice; B=contract month, D=Bloomberg ticker
    col_month = col('Ticker', 0)   # first occurrence  = col B = 'Jul26'
    col_bbg   = col('Ticker', 1)   # second occurrence = col D = 'CTN6 Comdty'

    options = {}

    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue

        market = _str(row, 0)
        if market == 'Holidays':
            break          # holiday table starts here — done reading options
        if market != 'CT':
            continue

        bbg_raw = _str(row, col_bbg) if col_bbg is not None else None
        if not bbg_raw or 'CT' not in bbg_raw:
            continue

        ticker = bbg_raw.replace(' Comdty', '').strip()   # 'CTN6 Comdty' -> 'CTN6'

        options[ticker] = {
            'contract_month': _str(row, col_month),                  # 'Jul26'
            'settle':         _num(row, col('Settle')),
            'atm_strike':     _num(row, col('Strike')),
            'call_ticker':    _str(row, col('Call Ticker')),
            'call_settle':    _num(row, col('Call Settle')),
            'put_ticker':     _str(row, col('Put Ticker')),
            'put_settle':     _num(row, col('Put Settle')),
            'strad_settle':   _num(row, col('Strad Settle')),
            'last':           _num(row, col('Last')),
            'current_strike': _num(row, col('Current Strike')),
            'call_value':     _num(row, col('Call Value')),
            'put_value':      _num(row, col('Put Value')),
            'strad_value':    _num(row, col('Strad Value')),
        }

    return options


# ── RTD sheet parser ──────────────────────────────────────────────────────────

def _parse_rtd(rows):
    """
    Cotton Straddles summary sheet.
    Columns: Contract, Strike, Value, ATM Vol, % CHG on Day,
             Settlement, Change, DTE, Expiration, Breakeven
    Contract codes are like 'N6', 'U6', 'Z6' (without 'CT' prefix).
    """
    if not rows:
        return {}

    # Find header row by scanning for 'Contract' in column A
    header_idx = None
    for i, row in enumerate(rows):
        if row and _hdr(row[0]) == 'Contract':
            header_idx = i
            break
    if header_idx is None:
        return {}

    headers = [_hdr(h) for h in rows[header_idx]]
    col     = _col_map(headers)

    straddles = {}

    for row in rows[header_idx + 1:]:
        if not row or row[0] is None:
            break

        code = _str(row, 0)       # 'N6', 'U6', 'Z6', 'H7' etc.
        if not code:
            break

        straddles[code] = {
            'strike':     _num(row, col.get('Strike',         1)),
            'value':      _num(row, col.get('Value',          2)),
            'atm_vol':    _num(row, col.get('ATM Vol',        3)),
            'pct_chg':    _num(row, col.get('% CHG on Day',   4)),
            'settlement': _num(row, col.get('Settlement',     5)),
            'change':     _num(row, col.get('Change',         6)),
            'dte':        _num(row, col.get('DTE',            7)),
            'expiration': row[col['Expiration']] if 'Expiration' in col else None,
        }

    return straddles


# ── all options BQL sheet parser ─────────────────────────────────────────────

def _parse_all_options(rows):
    """
    'all options' BQL sheet — cols 0-6 options, cols 11+ live futures.
    security_des format: 'CTN6C    65'  (split()[0] = 'CTN6C')
    expire_dt format:    '2026-06-12 00:00:00+00:00'  (ISO, take first 10 chars)
    Returns {
      'options': {ticker: {'expiry': 'YYYY-MM-DD', 'strikes': [...]}},
      'futures': {ticker: {'last': float, 'settle': float}}
    }
    """
    if not rows or not rows[0]:
        return {'options': {}, 'futures': {}}

    headers = [_hdr(h) for h in rows[0]]

    def _ci(keyword, skip=0):
        kw = keyword.lower()
        seen = 0
        for i, h in enumerate(headers):
            if kw in h.lower():
                if seen == skip:
                    return i
                seen += 1
        return None

    # Options columns (BQL query, left side)
    i_des      = _ci('security_des')
    i_strike   = _ci('strike_px')
    i_expiry   = _ci('expire_dt')
    i_bid      = _ci('px_bid')
    i_ask      = _ci('px_ask')
    i_opt_last = _ci('px_last')
    i_vol      = _ci('px_volume')

    # Futures columns (right side of same sheet, skip the options 'last' / 'ticker')
    i_fut_ticker = _ci('ticker', skip=1)   # second 'ticker' occurrence
    i_fut_last   = _ci('last',   skip=1)   # second 'last' occurrence (plain 'Last')
    i_fut_settle = _ci('settleprice')

    if i_des is None or i_strike is None:
        return {'options': {}, 'futures': {}}

    options = {}
    futures = {}

    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue

        # ── Live futures (right-side columns) ─────────────────────────────
        if i_fut_ticker is not None:
            ft_raw = _str(row, i_fut_ticker)
            if ft_raw and 'CT' in ft_raw and 'Comdty' in ft_raw:
                ftk = ft_raw.replace(' Comdty', '').strip()
                fl  = _num(row, i_fut_last)   if i_fut_last   is not None else None
                fs  = _num(row, i_fut_settle) if i_fut_settle is not None else None
                if ftk and (fl or fs):
                    futures[ftk] = {'last': fl, 'settle': fs}

        # ── Options (left-side columns) ───────────────────────────────────
        sec_des = _str(row, i_des)
        if not sec_des or not sec_des.startswith('CT'):
            continue

        # 'CTN6C    65' → split()[0] = 'CTN6C'
        first_token = sec_des.split()[0]
        if first_token.endswith('C'):
            ticker, pc = first_token[:-1], 'Call'
        elif first_token.endswith('P'):
            ticker, pc = first_token[:-1], 'Put'
        else:
            continue

        strike = _num(row, i_strike)
        if strike is None:
            continue

        # Expiry: '2026-06-12 00:00:00+00:00' — first 10 chars are YYYY-MM-DD
        expiry_str = None
        if i_expiry is not None and i_expiry < len(row):
            ev = row[i_expiry]
            if ev is not None:
                if hasattr(ev, 'strftime'):
                    expiry_str = ev.strftime('%Y-%m-%d')
                else:
                    s = str(ev).strip()
                    if len(s) >= 10 and s[4:5] == '-':
                        expiry_str = s[:10]
                    else:
                        for fmt in ('%m/%d/%Y', '%d/%m/%Y'):
                            try:
                                from datetime import datetime as _dt
                                expiry_str = _dt.strptime(s, fmt).strftime('%Y-%m-%d')
                                break
                            except ValueError:
                                continue

        bid  = _num(row, i_bid)
        ask  = _num(row, i_ask)
        last = _num(row, i_opt_last)
        vol  = _num(row, i_vol)

        mid = None
        if bid and ask and bid > 0.01 and ask > 0.01 and ask <= bid * 10:
            mid = (bid + ask) / 2.0
        if mid is None:
            mid = last

        if ticker not in options:
            options[ticker] = {'expiry': expiry_str, 'strikes': []}
        elif expiry_str and options[ticker]['expiry'] is None:
            options[ticker]['expiry'] = expiry_str

        options[ticker]['strikes'].append({
            'strike': strike, 'pc': pc,
            'bid': bid, 'ask': ask, 'mid': mid, 'last': last, 'vol': vol,
        })

    return {'options': options, 'futures': futures}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _hdr(v):
    """Normalise a header cell value to a stripped string, or ''."""
    if v is None:
        return ''
    return str(v).strip()


def _col_map(headers):
    """Build name->index dict, keeping the FIRST occurrence of duplicate names."""
    col = {}
    for i, h in enumerate(headers):
        if h and h not in col:
            col[h] = i
    return col


def _str(row, idx):
    """Return stripped string at row[idx], or None for errors/blanks."""
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    s = str(v).strip()
    return None if (not s or s.lower() in _BBG_ERRORS) else s


def _num(row, idx):
    """Return float at row[idx], or None for errors/blanks."""
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        # Bloomberg COM error codes arrive as large negative integers (e.g. -2146826273)
        if f < -1e8:
            return None
        return f
    s = str(v).strip().lower()
    if not s or s in _BBG_ERRORS:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _hv(row, idx):
    """Like _num but also returns None for 0.0 (Bloomberg placeholder when offline)."""
    v = _num(row, idx)
    return v if (v is not None and v > 0) else None
